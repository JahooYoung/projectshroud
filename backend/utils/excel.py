import os, re, string, random
import openpyxl
import datetime, pytz
import requests
from django.conf import settings
from backend.models import *
from backend.utils.email import send_activation_email


VERSION = '0.1.1'
REGISTRATION_API = 'http://localhost:8000/api/auth/registration/'
IMPORT_TEMPLATE_FILE_PATH = os.path.join(settings.FILES_DIR, 'template/Import_template.xlsx')
EXPORT_TEMPLATE_FILE_PATH = os.path.join(settings.FILES_DIR, 'template/Export_template.xlsx')
TEMP_FILES_DIR = os.path.join(settings.FILES_DIR, 'temp')
export_fields = {'user': ['real_name', 'mobile', 'email', 'is_activated'],
                 'status': ['approved', 'application_text'],
                 'transport': ['transport_type', 'transport_id', 'depart_station', 'depart_time',
                               'arrival_station', 'arrival_time', 'accommodation', 'other_detail']
                }
import_fields = {'user': ['real_name', 'mobile', 'email'],
                 'transport': ['transport_type', 'transport_id', 'depart_station', 'depart_time',
                               'arrival_station', 'arrival_time', 'accommodation', 'other_detail']
                }
MAGIC_STRING = 'tE5St3Sh1rOuD'
empty_re = r'^[\s]+$'
mobile_re = r'^(13[0-9]|14[579]|15[0-3,5-9]|16[6]|17[0135678]|18[0-9]|19[89])\d{8}$'
email_re = r'^[a-zA-Z0-9_-]+@[a-zA-Z0-9_-]+(\.[a-zA-Z0-9_-]+)+$'
EXPORT_START_ROW = 4
IMPORT_START_ROW = 9


def get_import_template():
    file = openpyxl.load_workbook(IMPORT_TEMPLATE_FILE_PATH)
    sheet = file['注册参会者']
    sheet['A1'].value = '%s VERSION %s' % (settings.SITE_OFFICIAL_NAME, VERSION)
    sheet['D1'].value = 'mgc-' + MAGIC_STRING
    file.save(IMPORT_TEMPLATE_FILE_PATH)
    return IMPORT_TEMPLATE_FILE_PATH


def fillrow(sheet, row, require_approve, ure):
    user = ure.user
    transport = ure.transport
    col = 1
    for field in export_fields['user']:
        sheet.cell(row=row, column=col).value = getattr(user, field)
        col += 1
    if require_approve:
        if ure.approved:
            sheet.cell(row=row, column=col).value = '审核通过'
        else:
            sheet.cell(row=row, column=col).value = '审核中'
            col += 1
            sheet.cell(row=row, column=col).value = ure.application_text
        col += 1
    else:
        if ure.approved:
            sheet.cell(row=row, column=col).value = '已注册'
        col += 1

    if transport is not None:
        for field in export_fields['transport']:
            val = getattr(transport, field)
            if isinstance(val, datetime.datetime):
                sheet.cell(row=row, column=col).number_format = 'yyyy/mm/dd hh:mm'
                val = val.astimezone(pytz.timezone(settings.TIME_ZONE))
                val = datetime.datetime(val.year, val.month, val.day, val.hour, val.minute)
            sheet.cell(row=row, column=col).value = val

            col += 1

    return 1


def export_excel(event):
    ure_list = UserRegisterEvent.objects.filter(event=event)
    file = openpyxl.load_workbook(EXPORT_TEMPLATE_FILE_PATH)

    sheet1 = file['已注册参会者']
    sheet2 = file['申请参会者']

    sheet2['A1'].value = sheet1['A1'].value = '%s VERSION %s' % (settings.SITE_OFFICIAL_NAME, VERSION)

    row1 = row2 = EXPORT_START_ROW

    for ure in ure_list:
        col = 0
        if ure.approved:
            row1 += fillrow(sheet1, row1, event.require_approve, ure)
        else:
            row2 += fillrow(sheet2, row2, event.require_approve, ure)

    row1 += 1
    sheet1.cell(row=row1, column=1).value = '总人数:'
    sheet1.cell(row=row1+1, column=1).value = row1 - 5

    row2 += 1
    sheet2.cell(row=row2, column=1).value = '总人数'
    sheet2.cell(row=row2+1, column=1).value = row2 - 5
    if not event.require_approve:
        del file['申请参会者']

    file_name = '%s_参会信息.xlsx' % re.sub(empty_re, '-', event.title)
    file_path = os.path.join(TEMP_FILES_DIR, file_name)
    file.save(file_path)

    return file_path, file_name


def register_new_user(data):
    def gen_password():
        passwd = ''
        passwd += ''.join(random.sample(string.digits, 2))
        passwd += ''.join(random.sample(string.ascii_letters, 7))
        passwd += ''.join(random.sample(string.digits, 1))
        return passwd

    passwd = gen_password()
    post_data = {}
    post_data['username'] = data['mobile']
    post_data['password1'] = post_data['password2'] = passwd
    post_data['real_name'] = data['real_name']
    post_data['email'] = data['email']

    return requests.post(REGISTRATION_API, data=post_data), passwd


def parserow(rdata, event):
    colnum = 0
    data = {}
    for field in import_fields['user']:
        data[field] = rdata[colnum].value
        colnum += 1
    data['mobile'] = str(data['mobile'])
    if data['mobile'] is None or not re.match(mobile_re, data['mobile']):
        return 0

    status = 1 # 0: failed; 1: User exists; 2: New user
    try:
        user = get_user_model().objects.get(mobile=data['mobile'])
    except get_user_model().DoesNotExist:
        status = 2
        if data['real_name'] is None or data['real_name'] == '' or re.match(empty_re, data['real_name']):
            return 0
        if data['email'] is None or not re.match(email_re, data['email']):
            return 0
        r, password = register_new_user(data)
        try:
            user = get_user_model().objects.get(mobile=data['mobile'])
        except get_user_model().DoesNotExist:
            return 0
        if r.status_code == 201:
            try:
                send_activation_email(user, event, password)
            except Exception:
                user.delete()
                return 0

    # Registration Success or already registered here
    if UserRegisterEvent.objects.filter(user=user, event=event).exists():
        return status
    event.attendee_count += 1
    event.save()
    for field in import_fields['transport']:
        data[field] = rdata[colnum].value if rdata[colnum].value is not None else ''
        # if isinstance(data[field], datetime.datetime):
        #     data[field] = data[field].replace(tzinfo=pytz.timezone(settings.TIME_ZONE))
        colnum += 1
    if data['transport_type'] not in ['Flight', '航班', 'Train', '列车', 'Other', '其他']:
        if data['transport_type'] == '航班':
            data['transport_type'] = 'Flight'
        if data['transport_type'] == '列车':
            data['transport_type'] = 'Train'
        if data['transport_type'] == '其他':
            data['transport_type'] = 'Other'
        UserRegisterEvent(user=user, event=event, transport=None, approved=True).save()
        return status
    if data['transport_id'] == '' or re.match(empty_re, data['transport_id']):
        UserRegisterEvent(user=user, event=event, transport=None, approved=True).save()
        return status
    if data['arrival_time'] == '':
        UserRegisterEvent(user=user, event=event, transport=None, approved=True).save()
        return status

    for field in import_fields['transport']:
        if field.endswith('time'):
            if isinstance(data[field], str):
                data[field] = datetime.datetime.strptime(data[field], '%Y/%m/%d %H:%M')
                # data[field] = data[field].replace(tzinfo=pytz.timezone(settings.TIME_ZONE))
    for field in import_fields['user']:
        del data[field]
    transport = Transport(user=user, event=event, **data)
    transport.save()

    UserRegisterEvent(user=user, event=event, transport=transport, approved=True).save()
    return status


def import_excel(event, file, callback):
    file_path = os.path.join(TEMP_FILES_DIR, '%s_upload.xlsx')
    with open(file_path, 'wb+') as destination:
        for chunk in file.chunks():
            destination.write(chunk)
    file = openpyxl.load_workbook(file_path)

    if '注册参会者' not in file:
        callback({'error': 'Table Name Altered!'})
        return
    sheet = file['注册参会者']
    if sheet['D1'].value != 'mgc-' + MAGIC_STRING:
        callback({'error': 'Magic String Altered!'})
        return

    rows = list(sheet.rows)
    total = sheet.max_row - (IMPORT_START_ROW - 1)
    suc, fail, user_count = 0, 0, 0

    for rownum in range(IMPORT_START_ROW-1, sheet.max_row):
        r = parserow(rows[rownum], event)
        if r != 0:
            suc += 1
            user_count += r - 1
        else:
            fail += 1
        callback({
            'finished': rownum == sheet.max_row - 1,
            'success_count': suc,
            'fail_count': fail,
            'user_count': user_count,
            'total': total
        })

    return suc, fail, user_count, total

